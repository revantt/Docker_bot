import sys
import time
import getpass
import urllib3
import xlsxwriter
import openpyxl as xl
from temporaryUtils import *
from openpyxl.styles import Font
from carrierData import CarrierData
from utils import getPrintableTable
from authentication import Authentication
import applicationsConstants as ApplicationsConstants

# To disable unverified HTTPS request is being made to host.
urllib3.disable_warnings()

# To not generate the bytecode i.e. .pyc files.
sys.dont_write_bytecode = True

def makeworkbook(filename):
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet("proposalsCount")
    headerFormat = workbook.add_format({'bold': True, 'fg_color': '3DD4F5', 'align': 'center', 'valign': 'vcenter'})
    worksheet.set_column('A:C', 45)
    activeInactiveHeader = ["ShipMethod", "Proposals Count"]
    worksheet.write_row(0, 0, activeInactiveHeader, headerFormat)
    workbook.close()

""" Fill workbook with given data """
def fillWorkbook(filename, carrier, shipmethods, shipmethodProposalsCount, lastFilledRowNumber):
    workbook = xl.load_workbook(filename)
    worksheet = workbook["proposalsCount"]
    worksheet.cell(lastFilledRowNumber, 1).value = carrier
    worksheet.cell(lastFilledRowNumber, 1).font = Font(bold=True)
    lastFilledRowNumber = lastFilledRowNumber + 1
    for shipmethod in shipmethods:
        worksheet.cell(lastFilledRowNumber, 1).value = shipmethod
        worksheet.cell(lastFilledRowNumber, 2).value = shipmethodProposalsCount[shipmethod]
        lastFilledRowNumber += 1
    workbook.save(filename)
    return lastFilledRowNumber

if __name__ == "__main__":
    printColoured(f"\nHello, you logged in as  {getpass.getuser()}", "green")
    printColoured(f"[INFO]: Authentication going on for user {getpass.getuser()} ...", "green")
    auth = Authentication(maxRedirects=15)
    auth.sentryAuthentication()
    printColoured("[INFO]: Sentry authentication done !!", "green")
    auth.midwayAuthentication()
    printColoured("[INFO]: Midway authentication done !!", "green")
    session = auth.session
    print("Note: Org list should be space separated, To best use of the script all" \
        " the configuration for carrier region and org should be correct")
    region = input("Please enter region: ")
    CarrierData("", region, "", False)
    orgs = input("Please enter org list: ")
    choice = 0
    while choice != '1' and choice != '2':
        print(f"Do you want proposal count for all carriers available in region: {region} \n1. Yes   2. No")
        choice = input("Please type 1 or 2: ")
    if int(choice) == 1:
        carrierData = CarrierData("", region, orgs, False)
        printColoured(f"[INFO]: Fetching carrier List for region {region} ...", "green")
        carrierList = getCarriersList(session, carrierData)
        printColoured(f"[INFO]: Number of carries available in region {region} are {len(carrierList)}", "green")
    else:
        carrier = input("Please enter carrier name: ")
        carrierData = CarrierData(carrier, region, orgs, False)
        carrierList = [carrierData.carrier]
    configElement = "TrackingIDMappings"
    filename = "proposalsCount_" + str(int(time.time())) +".xlsx"
    makeworkbook(filename)
    lastFilledRowNumber = 2
    for carrier in carrierList:
        carrierData = CarrierData(carrier, region, orgs, False)
        printColoured(f"[INFO]: Fetching ship methods for carrier {carrier} ...", "green")
        shipmethods = getShipmethods(session, carrierData)
        printColoured(f"[INFO]: Number of ship method available for carrier {carrier} are {len(shipmethods)}", "green")
        shipmethodProposalsCount = getShipmethodProposals(session, carrierData, shipmethods, configElement, False)
        lastFilledRowNumber = fillWorkbook(filename, carrier, shipmethods, shipmethodProposalsCount, lastFilledRowNumber + 1)
    print("\nWorkbook with name {} has been created\n".format(filename))
