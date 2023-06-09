import sys
import getpass
import urllib3
from utils import *
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import openpyxl as xl
from shutil import copyfile
from carrierData import CarrierData
from temporaryUtils import *
from openpyxl.styles import Alignment
from authentication import Authentication
import applicationsConstants as ApplicationsConstants
from shorttermutils import *

urllib3.disable_warnings()
sys.dont_write_bytecode = True

def fillworkbook(newmasterFile, carrierDetails):
    workbook = xl.load_workbook(newmasterFile)
    row = 8
    carrierDetailsWorksheet = workbook["CarrierStatus"]
    carrierDetailsWorksheet.cell(3, 1).value = carrierDetails["region"]
    carrierDetailsWorksheet.cell(3, 4).value = len(shipmethods)
    print('Filling data for:', carrierDetailsWorksheet.title)
    for shipmethod in carrierDetails["shipmethods"]:
        templates = carrierDetails["templates"]
        zpltemplates = ApplicationsConstants.NEW_LINE.join(templates[shipmethod]["ZPL"]) if "ZPL" in templates[shipmethod] else "-"
        pngtemplates = ApplicationsConstants.NEW_LINE.join(templates[shipmethod]["PNG"]) if "PNG" in templates[shipmethod] else  "-"
        kibanaStatus = carrierDetails["kibana_shipmethodstatus"][shipmethod]
        atropsStatus = carrierDetails["atrops_shipmethodstatus"][shipmethod]
        monitorportalStatus = carrierDetails["monitorportal_shipmethodstatus"][shipmethod]
        carrierDetailsWorksheet.cell(row, 1).value = shipmethod
        carrierDetailsWorksheet.cell(row, 2).value = zpltemplates
        carrierDetailsWorksheet.cell(row, 3).value = pngtemplates
        carrierDetailsWorksheet.cell(row, 4).value = kibanaStatus
        carrierDetailsWorksheet.cell(row, 5).value = atropsStatus
        carrierDetailsWorksheet.cell(row, 6).value = monitorportalStatus
        carrierDetailsWorksheet.cell(row, 8).value = carrierDetails["callRates"][shipmethod]["oneDay"]
        carrierDetailsWorksheet.cell(row, 9).value = carrierDetails["callRates"][shipmethod]["sevenDays"]
        row += 1

    sampleLabelsWorksheet = workbook["Shipmethods with Sample Labels"]
    print('Filling data for:', sampleLabelsWorksheet.title)
    row = 2
    for shipmethod in carrierDetails["shipmethods"]:
        if carrierDetails["imageStreams"][shipmethod] == ApplicationsConstants.EMPTY_STRING:
            continue
        templates = carrierDetails["templates"]
        zpltemplates = ApplicationsConstants.NEW_LINE.join(templates[shipmethod]["ZPL"]) if "ZPL" in templates[shipmethod] else "-"
        pngtemplates = ApplicationsConstants.NEW_LINE.join(templates[shipmethod]["PNG"]) if "PNG" in templates[shipmethod] else  "-"
        drivelink = "https://drive.corp.amazon.com/documents/" + getpass.getuser() + "@/" + driveDirName + shipmethod + ".jpg"
        containerId = carrierDetails["containerId"][shipmethod]
        sampleLabelsWorksheet.cell(row, 1).value = shipmethod
        sampleLabelsWorksheet.cell(row, 2).value = zpltemplates
        if pngtemplates != "-":
            sampleLabelsWorksheet.cell(row, 2).value += ApplicationsConstants.NEW_LINE + pngtemplates
        sampleLabelsWorksheet.cell(row, 3).value = '=HYPERLINK("{}", "{}")'.format(drivelink, shipmethod + ".jpg")
        sampleLabelsWorksheet.cell(row, 3).style= "Hyperlink"
        sampleLabelsWorksheet.cell(row, 4).value = containerId
        row += 1
    workbook.save(newmasterFile)

def fillSmDeconstructionDetail(filename, smDeconstructionData):
    print('Filling data for: SM-Deconstruction')
    workbook = xl.load_workbook(filename)
    worksheet = workbook["SM-Deconstruction"]
    smDeconstructionHeader = ["ShipMethod Name", "Region", "Service indicator", "Label Templates", "Weight",
    "Dimensions", "Package Value", "Other restrictions", "Atrops rate", "Atrops transit time", "Additional Information", "Queries"
    "Carrier Provided Service", "Value Added Services", "Why was it created as a seperate SM", "Was it due to some limitation in MLTS?", "comments"]
    row = 2
    for shipmethod in smDeconstructionData:
        shipmethodSMDeconstructionData = smDeconstructionData[shipmethod]
        for col in range(0, 9):
            if col in [2]:
                continue
            worksheet.cell(row, col+1).value = shipmethodSMDeconstructionData[smDeconstructionHeader[col]]
            if col == 0:
                worksheet.cell(row, col+1).value = '=HYPERLINK("{}", "{}")'.format(shipmethodSMDeconstructionData["Atrops Link"], shipmethod)
                worksheet.cell(row, col+1).style= "Hyperlink"
            worksheet.cell(row = 1, column = col + 1).alignment = Alignment(horizontal='center', vertical='center')
        row += 1
    workbook.save(filename)

def addG2S2AnalysisDetail(uniqueTemplates, templateG2S2Fields, templateShipmethodMap, CarrierG2S2Data, shipmethodG2S2Fields, newmasterFile):
    print('Adding G2S2 analysis data')
    workbook = xl.load_workbook(newmasterFile)
    for index, template in enumerate(uniqueTemplates):
        if len(template) > 31:
            templateG2S2Sheet = workbook.create_sheet(template[:31])
        else:
            templateG2S2Sheet = workbook.create_sheet(template)
        templateG2S2Sheet.cell(1, 1, "Template Name:")
        templateG2S2Sheet.cell(1, 2, template)
        templateG2S2Sheet.cell(1, 1).font = Font(bold=True)
        templateG2S2Sheet.cell(1, 1).alignment = Alignment(horizontal='center')
        templateG2S2Sheet.cell(1, 2).alignment = Alignment(horizontal='center')
        templateShipmethod = list(templateShipmethodMap[template])
        header = ["G2S2 field", "CarrierG2S2Data"]
        header.extend(templateShipmethod)
        for key, header_ in enumerate(header):
            templateG2S2Sheet.cell(3, key + 1, header_)
            templateG2S2Sheet.cell(3, key + 1).font = Font(bold=True)
            templateG2S2Sheet.cell(3, key + 1).alignment = Alignment(horizontal='center')
            templateG2S2Sheet.cell(3, key + 1).fill = PatternFill(start_color='3DD4F5',end_color='3DD4F5',fill_type='solid')

        for index_, g2s2field in enumerate(templateG2S2Fields[template]):
            templateG2S2Sheet.cell(index_ + 4, 1, g2s2field)
            templateG2S2Sheet.cell(index_ + 4, 1).alignment = Alignment(horizontal='center')
            if g2s2field in CarrierG2S2Data:
                try:
                    templateG2S2Sheet.cell(index_ + 4, 2, CarrierG2S2Data[g2s2field])
                except:
                    templateG2S2Sheet.cell(index_ + 4, 2, '\n'.join([f'{key}: {value}' for key, value in CarrierG2S2Data[g2s2field].items()]))
                templateG2S2Sheet.cell(index_ + 4, 2).alignment = Alignment(horizontal='center')

            for _index_, shipmethod in enumerate(templateShipmethod):
                if g2s2field in shipmethodG2S2Fields[shipmethod]:
                    try:
                        templateG2S2Sheet.cell(index_ + 4, _index_ + 3, shipmethodG2S2Fields[shipmethod][g2s2field])
                    except:
                        templateG2S2Sheet.cell(index_ + 4, _index_ + 3, \
                            '\n'.join([f'{key}: {value}' for key, value in shipmethodG2S2Fields[shipmethod][g2s2field].items()]))
                    templateG2S2Sheet.cell(index_ + 4, _index_ + 3).alignment = Alignment(horizontal='center')
    workbook.save(newmasterFile)

if __name__ == "__main__":
    printColoured(f"\nHello, you logged in as  {getpass.getuser()}", "green")
    printColoured(f"[INFO]: Authentication going on for user {getpass.getuser()} ...", "green")
    auth = Authentication(maxRedirects=20)
    auth.sentryAuthentication()
    printColoured("[INFO]: Sentry authentication done !!", "green")
    auth.midwayAuthentication()
    printColoured("[INFO]: Midway authentication done !!", "green")
    session = auth.session
    carrier, region, orgs = getCarrierRegionOrg()
    carrierData = CarrierData(carrier, region, orgs, False)
    printColoured("Fetching ship methods ...", "green")
    shipmethods = getShipmethodsFromExcel(session, carrierData)
    printColoured(f"[INFO]: Number of ship method available: {len(shipmethods)}", "green")
    if len(shipmethods) == 0:
        printColoured(f"[WARN]: Zero shipmethod found for carrier: {carrierData.carrier}", "yellow")
        sys.exit()
    newmasterFile = carrierData.carrier + ApplicationsConstants.masterFile
    copyfile(ApplicationsConstants.masterFile, newmasterFile)
    if len(shipmethods) > 0:
        username = getpass.getuser()
        printColoured("[INFO]: Getting templates from G2S2 ", "green")
        shipmethodTemplates, shipmethodShippingLabels = getShipmethodTemplates(session, carrierData, shipmethods)
        printColoured("[INFO]: Getting Active and Inactive status from Kibana", "green")
        kibanaShipmethodstatus = getActiveInactiveFromKibana(session, carrierData, shipmethods)
        printColoured("[INFO]: Getting Active and Inactive status from ATROPS", "green")
        atropsShipmethodstatus = getActiveInactiveFromAtrops(session, carrierData, shipmethods)
        printColoured("[INFO]: Getting Active and Inactive status from Monitorportal", "green")
        monitorportalShipmethodstatus = getActiveInactiveFromMonitorportal(session, carrierData, shipmethods)
        printColoured("[INFO]: Getting call rates from Monitorportal", "green")
        monitorportalShipmethodCallRate = getNDayCallRateFromMonitorportal(session, carrierData, shipmethods)
        printColoured("[INFO]: Getting Labels for all shipmethod ", "green")
        imageStreams, containerId = getLabelStreams(session, carrierData, shipmethods)
        driveDirName = carrier + "_LABELS/"
        makeNewDirInDrive(driveDirName)
        printColoured("[INFO]: Created new directory in drive with name " + driveDirName, "green")
        uploadImagesOnDrive(driveDirName, imageStreams)
        printColoured("[INFO]: All Label images are uploaded in drive folder " + driveDirName, "green")
        # printColoured("[INFO]: Getting SM Deconstruction data for all shipmethods ", "green")
        # smDeconstructionData = getSMDeconstructionData(session, carrierData, shipmethods)
        printColoured("[INFO]: Getting G2S2 fields from all templates ", "green")
        uniqueTemplates = getUniqueTemplatesFromCarrierTemplates(shipmethodTemplates)
        templateShipmethod = getMappedShipmethodToTemplates(shipmethodTemplates, uniqueTemplates)
        templateG2S2Fields = getG2S2FieldsForTemplates(session, carrierData, uniqueTemplates, templateShipmethod, shipmethodShippingLabels)
        shipmethodG2S2Fields = getG2S2FieldsForShipmethods(session, carrierData, shipmethods)
        CarrierG2S2Data = getG2S2FieldsForCarrier(session, carrierData)
        carrierDetails = { "region" : region, "shipmethods": shipmethods, "templates": shipmethodTemplates, "kibana_shipmethodstatus":\
            kibanaShipmethodstatus,  "atrops_shipmethodstatus": atropsShipmethodstatus, "monitorportal_shipmethodstatus":\
                  monitorportalShipmethodstatus, "imageStreams": imageStreams, "callRates": monitorportalShipmethodCallRate, "containerId": containerId}
        fillworkbook(newmasterFile, carrierDetails)
        # fillSmDeconstructionDetail(newmasterFile, smDeconstructionData)
        addG2S2AnalysisDetail(uniqueTemplates, templateG2S2Fields, templateShipmethod, CarrierG2S2Data, shipmethodG2S2Fields, newmasterFile)
    printColoured("\n[INFO]: Workbook with name " + newmasterFile + " has been created\n", "green")