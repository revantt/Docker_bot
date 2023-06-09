import sys
import time
import getpass
import urllib3
import xlsxwriter
import openpyxl as xl
from temporaryUtils import *
from openpyxl.styles import Font
from utils import getPrintableTable
from carrierData import CarrierData
from authentication import Authentication

""" To disable unverified HTTPS request is being made to host. """
urllib3.disable_warnings()

""" To not generate the bytecode i.e. .pyc files. """
sys.dont_write_bytecode = True

def makeworkbook(filename):
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet("ActiveInactive")
    headerFormat = workbook.add_format({'bold': True, 'fg_color': '3DD4F5', 'align': 'center', 'valign': 'vcenter'})
    secondheaderFormat = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter'})
    worksheet.set_column('A:H', 20)
    worksheet.merge_range(0, 1, 0, 4, 'Merged columns')
    activeInactiveHeader = ["ShipMethod", "Call Rates"]
    activeInactiveHeader1 = ["", "1 Day", "7 Days", "30 Days", "60 Days"]
    worksheet.write_row(0, 0, activeInactiveHeader, headerFormat)
    worksheet.write_row(0, 4, ["Call Rates"], headerFormat)
    worksheet.write_row(1, 0, activeInactiveHeader1, secondheaderFormat)
    workbook.close()

""" Make workbook with given data """
def fillWorkbook(filename, carrier, shipmethods, monitorportalShipmethodCallRate, lastFilledRowNumber):
    workbook = xl.load_workbook(filename)
    worksheet = workbook["ActiveInactive"]
    worksheet.cell(lastFilledRowNumber, 1).value = carrier
    worksheet.cell(lastFilledRowNumber, 1).font = Font(bold=True)
    lastFilledRowNumber = lastFilledRowNumber + 1
    for shipmethod in shipmethods:
        worksheet.cell(lastFilledRowNumber, 1).value = shipmethod
        worksheet.cell(lastFilledRowNumber, 2).value = monitorportalShipmethodCallRate[shipmethod]["oneDay"]
        worksheet.cell(lastFilledRowNumber, 3).value = monitorportalShipmethodCallRate[shipmethod]["sevenDays"]
        worksheet.cell(lastFilledRowNumber, 4).value = monitorportalShipmethodCallRate[shipmethod]["thirtyDays"]
        worksheet.cell(lastFilledRowNumber, 5).value = monitorportalShipmethodCallRate[shipmethod]["sixtyDays"]
        lastFilledRowNumber += 1
    workbook.save(filename)
    return lastFilledRowNumber

if __name__ == "__main__":
    print("\nHello, you logged in as:",  getpass.getuser())
    print("Authentication going on for user", getpass.getuser(), "...")
    auth = Authentication()
    auth.sentryAuthentication()
    print("Sentry authentication done !!")
    auth.midwayAuthentication()
    print("Midway authentication done !!")
    session = auth.session
    region = input("Please enter region: ")
    orgs = input("Please enter org list: ")
    carrierData = CarrierData("", region, orgs, False)
    print("Fetching carrier list ...")
    carriersList = getCarriersList(session, carrierData)
    print("Number of carriers available:", len(carriersList))
    filename = region + "_Volume_" + str(int(time.time())) +".xlsx"
    makeworkbook(filename)
    lastFilledRowNumber = 2
    for carrier in carriersList:
        carrierData = CarrierData(carrier, region, orgs, False)
        shipmethods = getShipmethods(session, carrierData)
        print("Number of shipmethods available for carrier {} is {}".format(carrier, len(shipmethods)))
        print("--------- Starting the MonitorPortal calls for volume ---------")
        monitorportalShipmethodCallRate = getNDayCallRateFromMonitorportal(session, carrierData, shipmethods)
        lastFilledRowNumber = fillWorkbook(filename, carrier, shipmethods, monitorportalShipmethodCallRate, lastFilledRowNumber + 1)
    print("\nWorkbook with name {} has been created\n".format(filename))
