import sys
import time
import getpass
import urllib3
import xlsxwriter
import openpyxl as xl
from temporaryUtils import *
from openpyxl.styles import Font
from carrierData import CarrierData
from utils import getPrintableTable
from authentication import Authentication
import applicationsConstants as ApplicationsConstants
from utils import getBeautifyHTML
from httpRequest import HttpRequest

# To disable unverified HTTPS request is being made to host.
urllib3.disable_warnings()

# To not generate the bytecode i.e. .pyc files.
sys.dont_write_bytecode = True

def makeworkbook(filename):
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet("DeliveryStatus")
    headerFormat = workbook.add_format({'bold': True, 'fg_color': '3DD4F5', 'align': 'center', 'valign': 'vcenter'})
    worksheet.set_column('A:E', 45)
    activeInactiveHeader = ["container", "trackingid", "orderid", "shipmentid", "deliverystatus", "esmmInfo"]
    worksheet.write_row(0, 0, activeInactiveHeader, headerFormat)
    workbook.close()

""" Fill workbook with given data """
def fillWorkbook(filename, container, trackingid, orderid, shipmentid, deliverystatus, esmmInfo, lastFilledRowNumber):
    workbook = xl.load_workbook(filename)
    worksheet = workbook["DeliveryStatus"]
    worksheet.cell(lastFilledRowNumber, 1).value = container
    worksheet.cell(lastFilledRowNumber, 2).value = trackingid
    worksheet.cell(lastFilledRowNumber, 3).value = orderid
    worksheet.cell(lastFilledRowNumber, 4).value = shipmentid
    worksheet.cell(lastFilledRowNumber, 5).value = deliverystatus
    worksheet.cell(lastFilledRowNumber, 6).value = esmmInfo
    lastFilledRowNumber = lastFilledRowNumber + 1
    workbook.save(filename)
    return lastFilledRowNumber

if __name__ == "__main__":
    printColoured(f"\nHello, you logged in as  {getpass.getuser()}", "green")
    printColoured(f"[INFO]: Authentication going on for user {getpass.getuser()} ...", "green")
    auth = Authentication(maxRedirects=15)
    auth.sentryAuthentication()
    printColoured("[INFO]: Sentry authentication done !!", "green")
    auth.midwayAuthentication()
    printColoured("[INFO]: Midway authentication done !!", "green")
    session = auth.session
    print("Note: Org list should be space separated, To best use of the script all" \
        " the configuration for carrier region and org should be correct")
    region = input("Please enter region: ")
    carrierData = CarrierData("", region, "", False)
    eagleeye = EagleEye(carrierData)
    filename = "eagleEyeInformation_" + str(int(time.time())) +".xlsx"
    makeworkbook(filename)
    lastFilledRowNumber = 2
    pasteId = input("Please enter the paste Id: ")
    printColoured("[INFO]: Parsing the paste data ", "green")
    requestResponse = HttpRequest(session, f"https://paste.amazon.com/show/{getpass.getuser()}/{pasteId}").getRequestResponse("GET")
    htmlData = getBeautifyHTML(requestResponse["data"])
    scannbleIds = [scannbleId for scannbleId in htmlData.find_all("textarea")[0].text.split("\n") if not scannbleId.strip() == ""]
    for scannbleId in scannbleIds:
        eagleEyeResponse = eagleeye.getEagleEyeResponse(session, scannbleId)
        container = eagleeye.getTCDAContainerId(eagleEyeResponse)
        deliveryStatus = eagleeye.getDeliveryStatus(eagleEyeResponse)
        trackingId = eagleeye.getTrackingId(eagleEyeResponse)
        orderId = eagleeye.getOrderId(eagleEyeResponse)
        shipmentid = eagleeye.getShipmentId(eagleEyeResponse)
        esmmInfo = eagleeye.getEsmmInfo(eagleEyeResponse)
        printColoured(f'[INFO]: Done with {lastFilledRowNumber-1}', "green")
        lastFilledRowNumber = fillWorkbook(filename, container, trackingId, orderId, shipmentid, deliveryStatus, esmmInfo, lastFilledRowNumber)
    printColoured("\nWorkbook with name {} has been created\n".format(filename), "green")
